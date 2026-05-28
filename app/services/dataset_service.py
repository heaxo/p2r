from __future__ import annotations

import io
import json
import math
import re
import shutil
import subprocess
import traceback
import uuid
import codecs
import xml.etree.ElementTree as ET
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

from fastapi import UploadFile
from loguru import logger

from app.core.algorithm import IMAGE_EXTS, json_safe
from app.errors import public_error_message, sanitize_public_error_message
from app.services.dxf_preview import create_dxf_preview
from app.services.dataset_store import DatasetStore
from app.services.lantek_registry import RegistryHelper
from app.services.measure_service import MeasureService


EXPERT_IMPORTERS = {"Procesos", "Masterlink"}


class DatasetNameConflict(ValueError):
    pass


class DatasetImportError(ValueError):
    pass


class DatasetService:
    """Dataset workflow built on top of the existing measuring service."""

    def __init__(self, measure_service: MeasureService) -> None:
        self.measure_service = measure_service
        self.settings = measure_service.settings
        self.store = DatasetStore(self.settings.task_db_path)
        self.store.mark_interrupted_work_failed()
        self._executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="plate-dataset-task")

    def default_name(self) -> str:
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def unique_name(self, base: str) -> str:
        base = (base or self.default_name()).strip() or self.default_name()
        if not self.store.name_exists(base):
            return base
        suffix = 2
        while self.store.name_exists(f"{base}_{suffix}"):
            suffix += 1
        return f"{base}_{suffix}"

    def create_dataset(self, name: str | None = None, *, source: str = "manual") -> Dict[str, Any]:
        clean_name = (name or "").strip()
        if clean_name:
            if self.store.name_exists(clean_name):
                raise DatasetNameConflict("数据集名称已存在")
        else:
            clean_name = self.unique_name(self.default_name())

        dataset = self.store.create_dataset(
            dataset_id=uuid.uuid4().hex,
            name=clean_name,
            source=source,
        )
        logger.info("Dataset created: id={}, name={}, source={}", dataset["id"], clean_name, source)
        return self.dataset_to_response(dataset)

    def list_datasets(self) -> list[Dict[str, Any]]:
        return [self.dataset_to_response(row) for row in self.store.list_datasets()]

    def get_dataset_detail(self, dataset_id: str) -> Dict[str, Any]:
        dataset = self.store.get_dataset(dataset_id)
        if dataset is None:
            raise KeyError(dataset_id)
        response = self.dataset_to_response(dataset)
        items = [self.item_to_response(row) for row in self.store.list_items(dataset_id)]
        response["items"] = items
        response["item_count"] = len(items)
        response["completed_count"] = sum(1 for item in items if item["status"] == "completed")
        response["failed_count"] = sum(1 for item in items if item["status"] == "failed")
        return response

    def update_import_settings(self, dataset_id: str, expert_importer: str) -> Dict[str, Any]:
        self._require_dataset(dataset_id)
        importer = self._normalize_expert_importer(expert_importer)
        self.store.update_dataset(dataset_id, expert_importer=importer)
        return self.get_dataset_detail(dataset_id)

    async def add_item(self, dataset_id: str, *, image: UploadFile, fields: Dict[str, Any]) -> Dict[str, Any]:
        self._require_dataset(dataset_id)
        image_path = await self.save_item_upload(dataset_id, image)
        item = self.store.add_item(
            id=uuid.uuid4().hex,
            dataset_id=dataset_id,
            row_order=self.store.next_item_order(dataset_id),
            image_path=str(image_path),
            original_filename=image.filename or image_path.name,
            **self.clean_item_fields(fields),
        )
        return self.item_to_response(item)

    async def update_item(
        self,
        dataset_id: str,
        item_id: str,
        *,
        image: UploadFile | None,
        fields: Dict[str, Any],
    ) -> Dict[str, Any]:
        self._require_dataset(dataset_id)
        current = self.store.get_item(item_id, dataset_id)
        if current is None:
            raise KeyError(item_id)

        updates = self.clean_item_fields(fields)
        if image is not None and image.filename:
            image_path = await self.save_item_upload(dataset_id, image)
            updates["image_path"] = str(image_path)
            updates["original_filename"] = image.filename or image_path.name

        updates.update(
            {
                "status": "pending",
                "progress": 0,
                "message": "",
                "result_json": None,
                "run_dir": None,
                "started_at": None,
                "finished_at": None,
            }
        )
        self.store.update_item(item_id, dataset_id, **updates)
        updated = self.store.get_item(item_id, dataset_id)
        if updated is None:
            raise KeyError(item_id)
        return self.item_to_response(updated)

    def delete_item(self, dataset_id: str, item_id: str) -> None:
        self._require_dataset(dataset_id)
        self.store.delete_item(item_id, dataset_id)

    def clear_item_recognition(self, dataset_id: str, item_id: str) -> Dict[str, Any]:
        dataset = self._require_dataset(dataset_id)
        if dataset.get("status") == "recognizing":
            raise ValueError("数据集正在识别中，不能清除识别状态")

        cleared = self.store.clear_item_recognition(item_id, dataset_id)
        if not cleared:
            raise ValueError("只有已识别的数据可以清除识别状态")
        return self.get_dataset_detail(dataset_id)

    def clear_dataset_recognition(self, dataset_id: str) -> Dict[str, Any]:
        dataset = self._require_dataset(dataset_id)
        if dataset.get("status") == "recognizing":
            raise ValueError("数据集正在识别中，不能清除识别状态")

        self.store.clear_dataset_recognition(dataset_id)
        return self.get_dataset_detail(dataset_id)

    def delete_dataset(self, dataset_id: str) -> None:
        dataset = self._require_dataset(dataset_id)
        if dataset.get("status") == "recognizing":
            raise ValueError("\u6570\u636e\u96c6\u6b63\u5728\u8bc6\u522b\u4e2d\uff0c\u4e0d\u80fd\u5220\u9664")

        items = self.store.list_items(dataset_id)
        self.store.delete_dataset(dataset_id)
        for item in items:
            self._safe_delete_output_path(item.get("image_path"))
            self._safe_delete_output_path(item.get("run_dir"))
        self._safe_delete_output_path(self.settings.output_root / "_datasets" / dataset_id)
        logger.info("Dataset deleted: id={}, name={}", dataset_id, dataset.get("name"))

    def copy_dataset(self, dataset_id: str, name: str | None = None) -> Dict[str, Any]:
        source = self.store.get_dataset(dataset_id)
        if source is None:
            raise KeyError(dataset_id)

        requested_name = (name or "").strip()
        if requested_name:
            if self.store.name_exists(requested_name):
                raise DatasetNameConflict("数据集名称已存在")
            new_name = requested_name
        else:
            new_name = self.unique_name(f"{source['name']}_副本")

        new_dataset_id = uuid.uuid4().hex
        dataset = self.store.create_dataset(
            dataset_id=new_dataset_id,
            name=new_name,
            source="copy",
            copied_from_id=dataset_id,
            expert_importer=source.get("expert_importer") or "Procesos",
        )
        for item in self.store.list_items(dataset_id):
            image_path = self._copy_existing_image(new_dataset_id, item.get("image_path"))
            self.store.add_item(
                id=uuid.uuid4().hex,
                dataset_id=new_dataset_id,
                row_order=int(item.get("row_order") or self.store.next_item_order(new_dataset_id)),
                image_path=str(image_path) if image_path else str(item.get("image_path") or ""),
                original_filename=item.get("original_filename") or (Path(str(image_path)).name if image_path else ""),
                plate_no=item.get("plate_no"),
                quantity=item.get("quantity"),
                material=item.get("material"),
                thickness_mm=item.get("thickness_mm"),
                dxf_target_size_1_mm=item.get("dxf_target_size_1_mm"),
                dxf_target_size_2_mm=item.get("dxf_target_size_2_mm"),
                dxf_target_x_mm=item.get("dxf_target_x_mm"),
                dxf_target_y_mm=item.get("dxf_target_y_mm"),
                paper_source=item.get("paper_source") or "yolo",
                a4_orientation=item.get("a4_orientation") or "auto",
                use_plate_perspective=bool(item.get("use_plate_perspective")),
                dxf_notch_fill_enabled=bool(item.get("dxf_notch_fill_enabled")),
                dxf_notch_fill_max_width_mm=item.get("dxf_notch_fill_max_width_mm"),
                dxf_notch_fill_max_depth_mm=item.get("dxf_notch_fill_max_depth_mm"),
            )
        logger.info("Dataset copied: source_id={}, new_id={}, name={}", dataset_id, new_dataset_id, new_name)
        return self.get_dataset_detail(dataset["id"])

    def enqueue_recognition(self, dataset_id: str, *, sam_model: str | None = None) -> Dict[str, Any]:
        dataset = self.store.get_dataset(dataset_id)
        if dataset is None:
            raise KeyError(dataset_id)
        if dataset.get("status") == "recognizing":
            raise ValueError("数据集正在识别中")

        selected_sam_model = self._as_text(sam_model) or None
        queued_count = self.store.begin_recognition(dataset_id)
        self._executor.submit(self._run_dataset, dataset_id, selected_sam_model)
        logger.info(
            "Dataset recognition queued: id={}, name={}, queued_count={}, sam_model={}",
            dataset_id,
            dataset.get("name"),
            queued_count,
            selected_sam_model or self.settings.sam_model,
        )
        return self.get_dataset_detail(dataset_id)

    def _run_dataset(self, dataset_id: str, sam_model: str | None = None) -> None:
        failed_count = 0
        last_error = None
        items = self.store.list_items(dataset_id)
        for item in items:
            if item.get("status") != "queued":
                continue

            item_id = item["id"]
            try:
                image_path = Path(str(item.get("image_path") or ""))
                if not image_path.exists():
                    raise FileNotFoundError(f"图片文件不存在：{image_path}")

                self.store.update_item(
                    item_id,
                    dataset_id,
                    status="processing",
                    progress=20,
                    message="正在识别",
                )
                args = self.measure_service.build_args(
                    image_path=image_path,
                    model_path=None,
                    sam_model=sam_model,
                    imgsz=None,
                    conf=None,
                    yolo_input_mode="canonical_path",
                    plate_class="plate",
                    paper_class="paper",
                    user_point_ratio=None,
                    paper_source=self._normalize_paper_source(item.get("paper_source")),
                    paper_sam2_yolo_fallback=False,
                    a4_orientation=self._normalize_a4_orientation(item.get("a4_orientation")),
                    perspective_source="plate" if bool(item.get("use_plate_perspective")) else "a4",
                    paper_points=None,
                    paper_rect_mode="approx_poly",
                    simplify_mm=3.0,
                    topdown_mm_per_px=2.0,
                    topdown_padding_mm=50.0,
                    enabled=True,
                    dxf_notch_fill_enabled=bool(item.get("dxf_notch_fill_enabled")),
                    dxf_notch_fill_max_width_mm=float(item.get("dxf_notch_fill_max_width_mm") or 80.0),
                    dxf_notch_fill_max_depth_mm=float(item.get("dxf_notch_fill_max_depth_mm") or 25.0),
                    dxf_target_size_1_mm=item.get("dxf_target_size_1_mm"),
                    dxf_target_size_2_mm=item.get("dxf_target_size_2_mm"),
                    dxf_target_x_mm=item.get("dxf_target_x_mm"),
                    dxf_target_y_mm=item.get("dxf_target_y_mm"),
                )
                self.store.update_item(item_id, dataset_id, progress=45, message="正在生成 DXF")
                result = self.measure_service.measure(args)
                result = self._rename_result_dxf(result, item)
                self.store.update_item(
                    item_id,
                    dataset_id,
                    status="completed",
                    progress=100,
                    message="识别完成",
                    result_json=json.dumps(json_safe(result), ensure_ascii=False),
                    run_dir=str(result.get("run_dir") or ""),
                )
                logger.info("Dataset item completed: dataset_id={}, item_id={}", dataset_id, item_id)
            except Exception as exc:
                failed_count += 1
                last_error = public_error_message(exc)
                print(traceback.format_exc())
                logger.exception("Dataset item failed: dataset_id={}, item_id={}, error={}", dataset_id, item_id, exc)
                self.store.update_item(
                    item_id,
                    dataset_id,
                    status="failed",
                    progress=100,
                    message=last_error,
                )

        self.store.finish_recognition(dataset_id, failed_count=failed_count, last_error=last_error)
        logger.info("Dataset recognition finished: id={}, failed_count={}", dataset_id, failed_count)

    async def import_excel(self, excel: UploadFile, name: str | None = None) -> Dict[str, Any]:
        if not excel.filename or not excel.filename.lower().endswith((".xlsx", ".xlsm")):
            raise DatasetImportError("请上传 .xlsx 或 .xlsm Excel 文件")

        content = await excel.read()
        await excel.close()
        rows = self._parse_excel_rows(content)
        if not rows:
            raise DatasetImportError("Excel 中没有可导入的数据")

        dataset_name = (name or "").strip()
        if dataset_name:
            if self.store.name_exists(dataset_name):
                raise DatasetNameConflict("数据集名称已存在")
        else:
            dataset_name = self.unique_name(Path(excel.filename).stem or self.default_name())

        dataset_id = uuid.uuid4().hex
        dataset_dir = self._dataset_dir(dataset_id)
        try:
            dataset = self.store.create_dataset(
                dataset_id=dataset_id,
                name=dataset_name,
                source="excel",
            )
            for index, row in enumerate(rows, start=1):
                source_path = self._resolve_import_image_path(row["image_path"])
                image_path = self.copy_image_from_path(dataset_id, source_path)
                self.store.add_item(
                    id=uuid.uuid4().hex,
                    dataset_id=dataset_id,
                    row_order=index,
                    image_path=str(image_path),
                    original_filename=source_path.name,
                    **self.clean_item_fields(row),
                )
        except Exception:
            if dataset_dir.exists():
                shutil.rmtree(dataset_dir, ignore_errors=True)
            try:
                self.store.delete_dataset(dataset_id)
            except Exception:
                pass
            raise

        logger.info("Dataset imported from Excel: id={}, filename={}, rows={}", dataset_id, excel.filename, len(rows))
        return self.get_dataset_detail(dataset["id"])

    def create_excel_template(self) -> bytes:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "数据集导入模板"

        headers = [
            ("图片地址", "必填。填写后端可访问的本机图片路径，桌面版建议填写绝对路径。"),
            ("板材编号", "可选。用于历史记录和结果区分。"),
            ("数量", "可选。填写大于 0 的整数。"),
            ("材质", "可选。"),
            ("厚度", "可选，单位 mm。"),
            ("尺寸1", "可选。不指定方向；尺寸1和尺寸2同时填写时，会按识别到的长短边自动匹配。"),
            ("尺寸2", "可选。不指定方向；尺寸1和尺寸2必须同时填写。"),
            ("x", "可选，单位 mm。明确指定 DXF X 方向目标尺寸。为空时优先使用尺寸1/尺寸2。"),
            ("y", "可选，单位 mm。明确指定 DXF Y 方向目标尺寸。为空时优先使用尺寸1/尺寸2。"),
            ("A4纸方向", "可选，默认 auto。用于指定A4纸在矫正后对应的毫米方向：auto 自动判断；x轴边297mm y轴边210mm；x轴边210mm y轴边297mm。拍摄角度导致系统判断错误时填写。"),
            ("是否启用钢板透视", "填写 是/否、true/false 或 1/0。启用后以钢板轮廓做透视矫正。"),
            ("是否启用夹钳修复", "填写 是/否、true/false 或 1/0。启用后修复夹钳造成的边缘凹陷。"),
            ("夹钳修复最大宽度mm", "可选，默认 80。夹钳凹陷最大宽度，单位 mm。"),
            ("夹钳修复最大深度mm", "可选，默认 25。夹钳凹陷最大深度，单位 mm。"),
        ]
        for index, (title, note) in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index, value=title)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2D579B")
            cell.comment = Comment(note, "p2r")
            sheet.column_dimensions[cell.column_letter].width = max(14, min(34, len(title) * 3))

        bool_validation = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
        sheet.add_data_validation(bool_validation)
        orientation_validation = DataValidation(
            type="list",
            formula1='"auto,x轴边297mm y轴边210mm,x轴边210mm y轴边297mm"',
            allow_blank=True,
        )
        sheet.add_data_validation(orientation_validation)
        orientation_validation.add("J2:J2000")
        bool_validation.add("K2:L2000")
        sheet.freeze_panes = "A2"

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def download_dataset_dxf_zip(self, dataset_id: str) -> tuple[bytes, str, int]:
        dataset = self._require_dataset(dataset_id)
        items = self.store.list_items(dataset_id)
        zip_buffer = io.BytesIO()
        added = 0
        used_names: set[str] = set()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            for item in items:
                if item.get("status") != "completed":
                    continue
                result = self._load_result(item)
                dxf_value = ((result or {}).get("paths") or {}).get("dxf")
                if not dxf_value:
                    continue
                dxf_path = Path(str(dxf_value))
                if not dxf_path.exists():
                    continue
                base_name = item.get("plate_no") or item.get("original_filename") or item.get("id") or "plate"
                safe_name = re.sub(r'[\\/:*?"<>|]+', "_", Path(str(base_name)).stem).strip(" .") or "plate"
                archive_name = f"{safe_name}.dxf"
                suffix = 2
                while archive_name in used_names:
                    archive_name = f"{safe_name}_{suffix}.dxf"
                    suffix += 1
                used_names.add(archive_name)
                archive.write(dxf_path, archive_name)
                added += 1

        if added == 0:
            raise FileNotFoundError("数据集中没有已完成的 DXF 文件")
        zip_buffer.seek(0)
        filename = f"{dataset['name']}_dxf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return zip_buffer.getvalue(), filename, added

    def import_to_expert(self, dataset_id: str) -> Dict[str, Any]:
        dataset, rows = self._expert_import_rows(dataset_id)
        importer = self._normalize_expert_importer(dataset.get("expert_importer") or "Procesos")
        import_dir = self._expert_import_dir(dataset)

        if importer == "Masterlink":
            generated = self._build_masterlink_import(import_dir, dataset, rows)
            batch_path = self._build_masterlink_batch(generated["xml_path"])
        else:
            generated = self._build_procesos_import(import_dir, dataset, rows)
            batch_path = self._build_procesos_batch(generated["prc_path"])

        result = self._run_import_batch(batch_path)
        return {
            "ok": True,
            "importer": importer,
            "dataset_id": dataset_id,
            "row_count": len(rows),
            "import_dir": str(import_dir),
            "batch_path": str(batch_path),
            "generated_files": {key: str(value) for key, value in generated.items()},
            "returncode": result.returncode,
            "message": "命令已执行完成，请打开套料软件确认是否成功导入",
        }

    def _expert_import_rows(self, dataset_id: str) -> tuple[Dict[str, Any], list[Dict[str, Any]]]:
        dataset = self._require_dataset(dataset_id)
        rows: list[Dict[str, Any]] = []
        for item in self.store.list_items(dataset_id):
            result = self._load_result(item)
            dxf_value = ((result or {}).get("paths") or {}).get("dxf")
            if not dxf_value:
                continue
            dxf_path = Path(str(dxf_value))
            if not dxf_path.exists() or not dxf_path.is_file():
                continue
            reference = (
                self._as_text(item.get("plate_no"))
                or self._safe_file_stem(item.get("original_filename"))
                or self._as_text(item.get("id"))
                or "plate"
            )
            rows.append(
                {
                    "reference": reference,
                    "material": self._as_text(item.get("material")),
                    "thickness": self._number_text(item.get("thickness_mm"), "0"),
                    "quantity": self._positive_int_text(item.get("quantity"), "1"),
                    "dxf_path": str(dxf_path.resolve()),
                }
            )

        if not rows:
            raise FileNotFoundError("数据集中没有可导入的已完成 DXF 文件")
        return dataset, rows

    def _build_procesos_import(self, import_dir: Path, dataset: Dict[str, Any], rows: list[Dict[str, Any]]) -> Dict[str, Path]:
        stem = self._safe_file_stem(dataset.get("name") or dataset.get("id") or "dataset")
        lst_path = import_dir / f"{stem}.LST"
        prc_path = import_dir / f"{stem}.prc"

        lines = []
        for row in rows:
            lines.append(
                f"{self._quote_lst(row['reference'])} "
                f"0 0 {row['thickness']} "
                f"{self._quote_lst(row['material'])} "
                f"{row['quantity']} 0 "
                f"{self._quote_lst('')} {self._quote_lst('')} {self._quote_lst('')} "
                f"2 {self._quote_lst(row['dxf_path'])}"
            )
        lst_path.write_text("\r\n".join(lines) + "\r\n", encoding="utf-8")
        prc_path.write_text(f"0 FILEPROLT 8.02\r\n39 1 {self._quote_lst(str(lst_path.resolve()))}\r\n", encoding="utf-8")
        return {"lst_path": lst_path, "prc_path": prc_path}

    def _build_masterlink_import(self, import_dir: Path, dataset: Dict[str, Any], rows: list[Dict[str, Any]]) -> Dict[str, Path]:
        stem = self._safe_file_stem(dataset.get("name") or dataset.get("id") or "dataset")
        xml_path = import_dir / f"{stem}.xml"
        root = ET.Element(
            "DATAEX",
            {
                "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
                "xmlns:xsd": "http://www.w3.org/2001/XMLSchema",
            },
        )

        for row in rows:
            product = ET.SubElement(root, "COMMAND", {"Name": "Import", "TblRef": "PRODUCTS", "Identifier": "0", "Task": "None"})
            self._xml_field(product, "Reference", row["reference"], "20")
            self._xml_field(product, "PCategory", "51", "70")
            self._xml_field(product, "Name", row["reference"], "20")
            self._xml_field(product, "Material", row["material"], "20")
            self._xml_field(product, "Thickness", row["thickness"], "100")
            self._xml_field(product, "Length", "0", "100")
            self._xml_field(product, "Width", "0", "100")
            self._xml_field(product, "CAMQuantity", row["quantity"], "100")
            self._xml_field(product, "Remnant", "1", "10")

            geometry = ET.SubElement(root, "COMMAND", {"Name": "Import", "TblRef": "IMPORTGEO", "Identifier": "0", "Task": "None"})
            self._xml_field(geometry, "Product", row["reference"], "20")
            self._xml_field(geometry, "GeometryPath", row["dxf_path"], "20")
            self._xml_field(geometry, "GeometryType", "dxf", "20")

        ET.indent(root, space="  ")
        content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        xml_path.write_bytes(codecs.BOM_UTF8 + content)
        return {"xml_path": xml_path}

    def _build_procesos_batch(self, prc_path: Path) -> Path:
        lantek_dir = self._require_lantek_install_dir()
        exe_path = lantek_dir / "Expert" / "Procesos.exe"
        if not exe_path.exists():
            raise FileNotFoundError(f"未找到 Procesos.exe：{exe_path}")
        batch_path = prc_path.with_name(f"prc-{uuid.uuid4().hex[:12]}.bat")
        batch_path.write_text(
            "\r\n".join(
                [
                    "@ECHO OFF",
                    "chcp 65001 >nul",
                    ":LOOP",
                    'tasklist | find /i "procesos.exe"',
                    "IF ERRORLEVEL 1 (",
                    "GOTO CONTINUE",
                    ") ELSE (",
                    "ECHO Procesos.exe is still running",
                    "Timeout /T 5 /Nobreak",
                    "GOTO LOOP",
                    ")",
                    ":CONTINUE",
                    f'START /W "" "{exe_path}" "{prc_path.resolve()}"',
                    "EXIT",
                ]
            ),
            encoding="utf-8",
        )
        return batch_path

    def _build_masterlink_batch(self, xml_path: Path) -> Path:
        lantek_dir = self._require_lantek_install_dir()
        exe_path = lantek_dir / "System" / "Common" / "XMLImporter.exe"
        if not exe_path.exists():
            raise FileNotFoundError(f"未找到 XMLImporter.exe：{exe_path}")
        batch_path = xml_path.with_name(f"xml-import-{uuid.uuid4().hex[:12]}.bat")
        batch_path.write_text(
            "\r\n".join(
                [
                    "@ECHO OFF",
                    "chcp 65001 >nul",
                    ":LOOP",
                    'tasklist | find /i "XMLImporter.exe"',
                    "IF ERRORLEVEL 1 (",
                    "GOTO CONTINUE",
                    ") ELSE (",
                    "ECHO XMLImporter.exe is still running",
                    "Timeout /T 5 /Nobreak",
                    "GOTO LOOP",
                    ")",
                    ":CONTINUE",
                    f'START /W "" "{exe_path}" -src "{xml_path.resolve()}"',
                    "EXIT",
                ]
            ),
            encoding="utf-8",
        )
        return batch_path

    def _run_import_batch(self, batch_path: Path) -> subprocess.CompletedProcess[str]:
        creationflags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0
        logger.info("Running Expert import batch: {}", batch_path)
        return subprocess.run(
            ["cmd.exe", "/c", str(batch_path)],
            cwd=str(batch_path.parent),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=creationflags,
            check=False,
        )

    def _require_lantek_install_dir(self) -> Path:
        install_dir = RegistryHelper.get_install_dir()
        if not install_dir:
            raise FileNotFoundError("未从注册表读取到 Lantek 安装目录：HKLM\\SOFTWARE\\WOW6432Node\\Lantek\\MainDir")
        return Path(install_dir)

    def _rename_result_dxf(self, result: Dict[str, Any], item: Dict[str, Any]) -> Dict[str, Any]:
        paths = dict(result.get("paths") or {})
        dxf_value = paths.get("dxf")
        if not dxf_value:
            return result

        dxf_path = Path(str(dxf_value))
        if not dxf_path.exists() or not dxf_path.is_file():
            return result

        base_name = item.get("plate_no") or item.get("original_filename") or dxf_path.stem
        safe_name = self._safe_file_stem(base_name)
        target = dxf_path.with_name(f"{safe_name}.dxf")
        if target.resolve() != dxf_path.resolve():
            suffix = 2
            while target.exists():
                target = dxf_path.with_name(f"{safe_name}_{suffix}.dxf")
                suffix += 1
            try:
                dxf_path.replace(target)
            except Exception as exc:
                logger.warning("Dataset DXF rename failed: source={}, target={}, error={}", dxf_path, target, exc)
                target = dxf_path

        paths["dxf"] = str(target)
        try:
            preview_path = create_dxf_preview(str(target))
            if preview_path:
                paths["dxf_preview"] = str(preview_path)
        except Exception as exc:
            logger.warning("Dataset renamed DXF preview generation failed: dxf_path={}, error={}", target, exc)

        updated = dict(result)
        updated["paths"] = paths
        updated["urls"] = {key: url for key, value in paths.items() if (url := self.measure_service._path_to_url(str(value)))}
        return updated

    async def save_item_upload(self, dataset_id: str, image: UploadFile) -> Path:
        if image is None or not image.filename:
            raise ValueError("请上传图片")
        suffix = self._safe_image_suffix(image.filename)
        target = self._unique_dataset_image_path(dataset_id, Path(image.filename).stem, suffix)
        max_bytes = self.settings.max_upload_mb * 1024 * 1024
        written = 0
        with target.open("wb") as f:
            while True:
                chunk = await image.read(1024 * 1024)
                if not chunk:
                    break
                written += len(chunk)
                if written > max_bytes:
                    f.close()
                    target.unlink(missing_ok=True)
                    raise ValueError(f"上传图片超过限制：{self.settings.max_upload_mb}MB")
                f.write(chunk)
        await image.close()
        return target

    def copy_image_from_path(self, dataset_id: str, source_path: Path) -> Path:
        if not source_path.exists() or not source_path.is_file():
            raise FileNotFoundError(f"图片文件不存在：{source_path}")
        suffix = self._safe_image_suffix(source_path.name)
        target = self._unique_dataset_image_path(dataset_id, source_path.stem, suffix)
        shutil.copy2(source_path, target)
        return target

    def clean_item_fields(self, fields: Dict[str, Any]) -> Dict[str, Any]:
        size_1 = self._optional_positive_float(fields.get("dxf_target_size_1_mm"), "尺寸1")
        size_2 = self._optional_positive_float(fields.get("dxf_target_size_2_mm"), "尺寸2")
        if (size_1 is None) != (size_2 is None):
            raise ValueError("尺寸1和尺寸2必须同时填写，或同时留空")

        return {
            "plate_no": self._as_text(fields.get("plate_no")),
            "quantity": self._optional_positive_int(fields.get("quantity"), "数量"),
            "material": self._as_text(fields.get("material")),
            "thickness_mm": self._optional_positive_float(fields.get("thickness_mm"), "厚度"),
            "dxf_target_size_1_mm": size_1,
            "dxf_target_size_2_mm": size_2,
            "dxf_target_x_mm": self._optional_positive_float(fields.get("dxf_target_x_mm"), "x"),
            "dxf_target_y_mm": self._optional_positive_float(fields.get("dxf_target_y_mm"), "y"),
            "paper_source": self._normalize_paper_source(fields.get("paper_source")),
            "a4_orientation": self._normalize_a4_orientation(fields.get("a4_orientation")),
            "use_plate_perspective": self._parse_bool(fields.get("use_plate_perspective")),
            "dxf_notch_fill_enabled": self._parse_bool(fields.get("dxf_notch_fill_enabled")),
            "dxf_notch_fill_max_width_mm": self._optional_positive_float(
                fields.get("dxf_notch_fill_max_width_mm"),
                "夹钳修复最大宽度",
                default=80.0,
            ),
            "dxf_notch_fill_max_depth_mm": self._optional_positive_float(
                fields.get("dxf_notch_fill_max_depth_mm"),
                "夹钳修复最大深度",
                default=25.0,
            ),
        }

    def dataset_to_response(self, row: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "id": row.get("id"),
            "name": row.get("name"),
            "status": row.get("status"),
            "source": row.get("source"),
            "expert_importer": row.get("expert_importer") or "Procesos",
            "copied_from_id": row.get("copied_from_id"),
            "last_error": sanitize_public_error_message(row.get("last_error") or "", fallback=""),
            "created_at": row.get("created_at"),
            "updated_at": row.get("updated_at"),
            "recognized_at": row.get("recognized_at"),
            "recognition_started_at": row.get("recognition_started_at"),
            "recognition_finished_at": row.get("recognition_finished_at"),
            "item_count": int(row.get("item_count") or 0),
            "completed_count": int(row.get("completed_count") or 0),
            "failed_count": int(row.get("failed_count") or 0),
        }

    def item_to_response(self, row: Dict[str, Any]) -> Dict[str, Any]:
        result = self._load_result(row)
        urls = dict((result or {}).get("urls") or {})
        paths = dict((result or {}).get("paths") or {})
        dxf_path = paths.get("dxf")
        if dxf_path and not paths.get("dxf_preview"):
            try:
                preview_path = create_dxf_preview(str(dxf_path))
                if preview_path:
                    paths["dxf_preview"] = str(preview_path)
            except Exception as exc:
                logger.warning("Dataset DXF preview generation failed: item_id={}, dxf_path={}, error={}", row.get("id"), dxf_path, exc)
        urls.update({key: url for key, value in paths.items() if (url := self.measure_service._path_to_url(str(value)))})
        image_url = self.measure_service._path_to_url(str(row.get("image_path") or ""))
        if image_url:
            urls["image"] = image_url

        return {
            "id": row.get("id"),
            "dataset_id": row.get("dataset_id"),
            "row_order": int(row.get("row_order") or 0),
            "image_path": row.get("image_path"),
            "original_filename": row.get("original_filename"),
            "plate_no": row.get("plate_no"),
            "quantity": row.get("quantity"),
            "material": row.get("material"),
            "thickness_mm": row.get("thickness_mm"),
            "dxf_target_size_1_mm": row.get("dxf_target_size_1_mm"),
            "dxf_target_size_2_mm": row.get("dxf_target_size_2_mm"),
            "dxf_target_x_mm": row.get("dxf_target_x_mm"),
            "dxf_target_y_mm": row.get("dxf_target_y_mm"),
            "paper_source": self._normalize_paper_source(row.get("paper_source")),
            "a4_orientation": self._normalize_a4_orientation(row.get("a4_orientation")),
            "use_plate_perspective": bool(row.get("use_plate_perspective")),
            "dxf_notch_fill_enabled": bool(row.get("dxf_notch_fill_enabled")),
            "dxf_notch_fill_max_width_mm": row.get("dxf_notch_fill_max_width_mm"),
            "dxf_notch_fill_max_depth_mm": row.get("dxf_notch_fill_max_depth_mm"),
            "status": row.get("status"),
            "progress": int(row.get("progress") or 0),
            "message": sanitize_public_error_message(row.get("message") or "", fallback=""),
            "created_at": row.get("created_at"),
            "updated_at": row.get("updated_at"),
            "started_at": row.get("started_at"),
            "finished_at": row.get("finished_at"),
            "run_dir": row.get("run_dir"),
            "urls": urls,
            "paths": paths,
            "plate_dimensions": (result or {}).get("plate_dimensions", {}),
            "dxf_target_size": (result or {}).get("dxf_target_size", {}),
            "detected_plate_dimensions": (result or {}).get("detected_plate_dimensions", {}),
            "dxf_geometry": (result or {}).get("dxf_geometry", {}),
            "dxf_postprocess": (result or {}).get("dxf_postprocess", {}),
            "result": result,
        }

    def _parse_excel_rows(self, content: bytes) -> list[Dict[str, Any]]:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [self._normalize_header(cell.value) for cell in sheet[1]]
        aliases = {
            "图片地址": "image_path",
            "图片路径": "image_path",
            "板材编号": "plate_no",
            "编号": "plate_no",
            "数量": "quantity",
            "材质": "material",
            "厚度": "thickness_mm",
            "尺寸1": "dxf_target_size_1_mm",
            "尺寸2": "dxf_target_size_2_mm",
            "x": "dxf_target_x_mm",
            "y": "dxf_target_y_mm",
            "a4纸方向": "a4_orientation",
            "a4方向": "a4_orientation",
            "纸张方向": "a4_orientation",
            "纸张毫米方向": "a4_orientation",
            "是否启用钢板透视": "use_plate_perspective",
            "钢板透视": "use_plate_perspective",
            "是否启用夹钳修复": "dxf_notch_fill_enabled",
            "夹钳修复": "dxf_notch_fill_enabled",
            "夹钳修复最大宽度mm": "dxf_notch_fill_max_width_mm",
            "夹钳修复最大深度mm": "dxf_notch_fill_max_depth_mm",
        }
        mapped = [aliases.get(header, header) for header in headers]
        if "image_path" not in mapped:
            raise DatasetImportError("Excel 模板缺少“图片地址”列")

        rows: list[Dict[str, Any]] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            data = {mapped[index]: value for index, value in enumerate(row) if index < len(mapped)}
            image_path = self._as_text(data.get("image_path"))
            if not image_path:
                raise DatasetImportError(f"第 {row_index} 行缺少图片地址")
            data["image_path"] = image_path
            self.clean_item_fields(data)
            rows.append(data)
        return rows

    def _resolve_import_image_path(self, value: str) -> Path:
        path = Path(value).expanduser()
        if not path.is_absolute():
            path = Path.cwd() / path
        return path.resolve()

    def _copy_existing_image(self, dataset_id: str, path_value: Any) -> Path | None:
        if not path_value:
            return None
        source = Path(str(path_value))
        if not source.exists() or not source.is_file():
            return None
        try:
            return self.copy_image_from_path(dataset_id, source)
        except Exception as exc:
            logger.warning("Failed to copy dataset item image: source={}, error={}", source, exc)
            return None

    def _safe_delete_output_path(self, path_value: Any) -> None:
        if not path_value:
            return

        try:
            root = self.settings.output_root.resolve()
            path = Path(str(path_value)).resolve()
            path.relative_to(root)
        except Exception:
            logger.warning("Skip deleting dataset path outside output_root: {}", path_value)
            return

        try:
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)
            elif path.exists():
                path.unlink(missing_ok=True)
        except Exception as exc:
            logger.warning("Failed to delete dataset path: path={}, error={}", path, exc)

    def _dataset_dir(self, dataset_id: str) -> Path:
        path = self.settings.output_root / "_datasets" / dataset_id
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _expert_import_dir(self, dataset: Dict[str, Any]) -> Path:
        dataset_id = str(dataset.get("id") or uuid.uuid4().hex)
        path = self.settings.output_root / "_expert_imports" / dataset_id / datetime.now().strftime("%Y%m%d_%H%M%S")
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _unique_dataset_image_path(self, dataset_id: str, stem: str, suffix: str) -> Path:
        images_dir = self._dataset_dir(dataset_id) / "images"
        images_dir.mkdir(parents=True, exist_ok=True)
        safe_stem = re.sub(r'[\\/:*?"<>|\s]+', "_", stem or "image").strip("._") or "image"
        return images_dir / f"{safe_stem}_{uuid.uuid4().hex[:8]}{suffix}"

    def _safe_image_suffix(self, filename: str) -> str:
        suffix = Path(filename or "").suffix.lower()
        if suffix not in IMAGE_EXTS:
            raise ValueError(f"不支持的图片格式：{suffix or 'unknown'}")
        return suffix

    def _safe_file_stem(self, value: Any) -> str:
        stem = Path(self._as_text(value) or "plate").stem
        return re.sub(r'[\\/:*?"<>|\s]+', "_", stem).strip(" ._") or "plate"

    def _normalize_expert_importer(self, value: Any) -> str:
        importer = self._as_text(value) or "Procesos"
        if importer.lower() == "masterlink":
            return "Masterlink"
        if importer.lower() == "procesos":
            return "Procesos"
        raise ValueError("导入器只能选择 Procesos 或 Masterlink")

    def _quote_lst(self, value: Any) -> str:
        text = self._as_text(value).replace('"', '""')
        return f'"{text}"'

    def _xml_field(self, parent: ET.Element, ref: str, value: Any, field_type: str) -> None:
        ET.SubElement(
            parent,
            "FIELD",
            {
                "FldRef": ref,
                "FldValue": self._as_text(value),
                "FldType": field_type,
            },
        )

    def _number_text(self, value: Any, default: str) -> str:
        if value is None or value == "":
            return default
        try:
            number = float(value)
        except (TypeError, ValueError):
            return default
        if not math.isfinite(number):
            return default
        return str(int(number)) if number.is_integer() else f"{number:.6f}".rstrip("0").rstrip(".")

    def _positive_int_text(self, value: Any, default: str) -> str:
        if value is None or value == "":
            return default
        try:
            number = int(float(value))
        except (TypeError, ValueError):
            return default
        return str(number) if number > 0 else default

    def _require_dataset(self, dataset_id: str) -> Dict[str, Any]:
        dataset = self.store.get_dataset(dataset_id)
        if dataset is None:
            raise KeyError(dataset_id)
        return dataset

    def _load_result(self, row: Dict[str, Any]) -> Dict[str, Any] | None:
        raw = row.get("result_json")
        if not raw:
            return None
        try:
            parsed = json.loads(raw)
            return parsed if isinstance(parsed, dict) else None
        except Exception:
            return None

    def _normalize_header(self, value: Any) -> str:
        return self._as_text(value).replace(" ", "").replace("_", "").lower()

    def _as_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _parse_bool(self, value: Any) -> bool:
        if isinstance(value, bool):
            return value
        text = self._as_text(value).lower()
        if not text:
            return False
        return text in {"1", "true", "yes", "y", "on", "是", "启用", "开启"}

    def _normalize_paper_source(self, value: Any) -> str:
        text = self._as_text(value).lower()
        if not text:
            return "yolo"
        if text in {"yolo", "training", "train", "trained", "dataset"}:
            return "yolo"
        if text in {"sam2", "standard", "std"}:
            return "sam2"
        raise ValueError("A4纸来源参数不正确")

    def _normalize_a4_orientation(self, value: Any) -> str:
        text = self._as_text(value).lower()
        if not text:
            return "auto"
        compact = text.replace(" ", "").replace("_", "").replace("-", "")
        if compact in {"auto", "自动", "自动判断", "自动识别"}:
            return "auto"
        if compact in {
            "landscape",
            "横向",
            "横版",
            "x297y210",
            "x轴边297mmy轴边210mm",
            "x边297mmy边210mm",
            "宽297高210",
            "297x210",
            "297*210",
        }:
            return "landscape"
        if compact in {
            "portrait",
            "纵向",
            "竖向",
            "竖版",
            "x210y297",
            "x轴边210mmy轴边297mm",
            "x边210mmy边297mm",
            "宽210高297",
            "210x297",
            "210*297",
        }:
            return "portrait"
        raise ValueError("A4纸方向参数不正确，请填写 auto、x轴边297mm y轴边210mm 或 x轴边210mm y轴边297mm")

    def _optional_positive_float(self, value: Any, field_name: str, default: float | None = None) -> float | None:
        if value is None:
            return default
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return default
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{field_name} 必须是大于 0 的数字") from exc
        if not math.isfinite(number) or number <= 0:
            raise ValueError(f"{field_name} 必须是大于 0 的数字")
        return number

    def _optional_positive_int(self, value: Any, field_name: str, default: int | None = None) -> int | None:
        if value is None:
            return default
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return default
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{field_name} 必须是大于 0 的整数") from exc
        if not math.isfinite(number) or number <= 0 or not number.is_integer():
            raise ValueError(f"{field_name} 必须是大于 0 的整数")
        return int(number)
